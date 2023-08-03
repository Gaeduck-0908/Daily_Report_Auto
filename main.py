import os
import datetime
from shutil import copyfile
import openpyxl

today = datetime.datetime.today()

# 원본 파일 경로 및 파일명 설정
src_file = "D:\\01.ETC\\Excel\\일일보고\\일일보고_김태한_000000.xlsx"

# 목적 파일 경로 및 파일명 설정
dest_file = "D:\\01.ETC\\Excel\\일일보고\\일일보고_김태한_{:02d}{:02d}{:02d}.xlsx".format(today.year % 100, today.month, today.day)

# 파일 복사
copyfile(src_file, dest_file)

# 엑셀 파일 수정
wb = openpyxl.load_workbook(dest_file)
ws = wb.active

# C2, D6, D13 셀에 오늘 날짜 업데이트
ws['C2'] = today.strftime("%y.%m.%d")
ws['D6'] = today.strftime("%m.%d")
ws['D13'] = (today+datetime.timedelta(1)).strftime("%m.%d")

# 엑셀 파일 저장
wb.save(dest_file)
wb.close()

print("작업 완료")
